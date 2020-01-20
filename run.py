import glob, os
import os.path
import datetime
import shutil
import platform
import io
import json
import datetime 
from dateutil.parser import parse

from openpyxl import load_workbook
import config as cfg

def tratadatas(arrDatas):
    menorData = None
    maiorData = None
    delta = 0
    lstdatas = []
    try:
        datas = arrDatas.split(' ')
        for d in datas:
            try:
                data = datetime.datetime.strptime(d, '%d/%m/%Y')
                lstdatas.append(data)
                lstdatas.sort()
            except:
                continue
        menorData = lstdatas[0]    
        maiorData = len(lstdatas) - 1   
        d1 = lstdatas[0].strftime('%d/%m/%Y')
        d2 = lstdatas[maiorData].strftime('%d/%m/%Y')
        delta  = parse(d2) - parse(d1)
        return lstdatas[0].strftime('%d/%m/%Y'),lstdatas[maiorData].strftime('%d/%m/%Y'),delta.days
    except:
        return '','',0
    # return lstdatas[0].strftime('%d/%m/%Y'),lstdatas[maiorData].strftime('%d/%m/%Y'),delta.days


def convertDataParse(strdata):
    data = strdata
    try:
        if not isinstance(data, datetime.datetime):
            dataconvertida = parse(data)
            data = dataconvertida.strftime('%d/%m/%Y')
        else:
            data = dataconvertida.strftime('%d/%m/%Y')
    except:
        data = strdata
    return data

def salvaProjetos(lista):
    with io.open('projetostodos.json', 'w', encoding='utf8', errors='ignore') as outfile:  
        json.dump(lista, outfile, ensure_ascii=False)

def readFilesFromSource():
    # os.chdir("pendentes")
    os.chdir(cfg.origemArquivosExcel)

    projetos = {}
    listaDeProjetos = []

    for i in os.listdir(os.getcwd()):
        print (os.path.join(os.getcwd(), i))
        print(i)
        if os.path.isfile(i):
            # if i.endswith(".xls") or i.endswith(".xlsx") or i.endswith(".xlsm"):
            if i.endswith(".xlsx"):
                DataCriacaoDoArquivo = creation_date(os.path.join(os.getcwd(), i))
                print ("Data criação do Arquivo: {}".format(datetime.datetime.fromtimestamp(DataCriacaoDoArquivo)))
                NomeDoArquivo = (os.path.join(os.getcwd(), i))
                NomeDoArquivo1 = (os.path.basename(NomeDoArquivo))
                print("Filename: {}".format(NomeDoArquivo))
                print("Filename with path: {}".format(NomeDoArquivo1))
                wb = load_workbook(os.path.join(os.getcwd(), i))
                ws_Projetos = wb['Projetos']
                total_linhas = ws_Projetos.max_row
                print("Total de Linhas na aba projetos: {}".format(total_linhas))
                
                for linha in range(cfg.linhaInicial, total_linhas+1):
                    projetos = {}
                    if ws_Projetos.cell(row=linha, column=cfg.excelColumn['Projeto']).value == None:
                        break
                    else:
                        projetos.update({'Nome do Arquivo': NomeDoArquivo1})
                        projetos.update({'Projeto': ws_Projetos.cell(row=linha, column=cfg.excelColumn['Projeto']).value})
                        projetos.update({'VP': ws_Projetos.cell(row=linha, column=cfg.excelColumn['VP']).value})
                        projetos.update({'Grupo Solucionador': ws_Projetos.cell(row=linha, column=cfg.excelColumn['Grupo Solucionador']).value})
                        projetos.update({'Programa': ws_Projetos.cell(row=linha, column=cfg.excelColumn['Programa']).value})
                        projetos.update({'Porte': ws_Projetos.cell(row=linha, column=cfg.excelColumn['Porte']).value})
                        projetos.update({'Status': ws_Projetos.cell(row=linha, column=cfg.excelColumn['Status']).value})
                        projetos.update({'Data de Início do Projeto': convertDataParse(str(ws_Projetos.cell(row=linha, column=cfg.excelColumn['Data de Início do Projeto']).value))})
                        
                        # dtinicioProjeto = str(ws_Projetos.cell(row=linha, column=cfg.excelColumn['Data de Início do Projeto']).value)

                        projetos.update({'Programa': ws_Projetos.cell(row=linha, column=cfg.excelColumn['Programa']).value})
                        
                        inicioEF = convertDataParse(str(ws_Projetos.cell(row=linha, column=cfg.excelColumn['Inicio EF']).value))
                        projetos.update({'Inicio EF': inicioEF})
                        terminioEF = convertDataParse(str(ws_Projetos.cell(row=linha, column=cfg.excelColumn['Termino EF']).value))
                        projetos.update({'Termino EF': terminioEF})
                        
                        menorData, maiorData, delta = tratadatas(inicioEF)
                        projetos.update({'Real Inicio EF': maiorData})
                        projetos.update({'Delta Inicio EF': delta})
                        
                        menorData, maiorData, delta = tratadatas(terminioEF)
                        projetos.update({'Real Termino EF': terminioEF})
                        projetos.update({'Delta Termino EF': delta})
                        projetos.update({'Complitude EF': str(ws_Projetos.cell(row=linha, column=cfg.excelColumn['Complitude EF']).value)})
                        

                        inicioET = convertDataParse(str(ws_Projetos.cell(row=linha, column=cfg.excelColumn['Início ET']).value))
                        projetos.update({'Início ET': inicioET})
                        terminoET = convertDataParse(str(ws_Projetos.cell(row=linha, column=cfg.excelColumn['Término ET']).value))
                        projetos.update({'Término ET': terminoET})
                        projetos.update({'Complitude ET': str(ws_Projetos.cell(row=linha, column=cfg.excelColumn['Complitude ET']).value)})
                        
                        menorData, maiorData, delta = tratadatas(inicioET)
                        projetos.update({'Real Início ET': maiorData})
                        projetos.update({'Delta Início ET': delta})
                        
                        menorData, maiorData, delta = tratadatas(terminoET)
                        projetos.update({'Real Término ET': maiorData})
                        projetos.update({'Delta Término ET': delta})

                        inicioDesenv = convertDataParse(str(ws_Projetos.cell(row=linha, column=cfg.excelColumn['Início do Desenv.']).value))
                        projetos.update({'Início do Desenv.': inicioDesenv})
                        terminoDesenv = convertDataParse(str(ws_Projetos.cell(row=linha, column=cfg.excelColumn['Término do Desenv.']).value))
                        projetos.update({'Término do Desenv.': terminoDesenv})
                        
                        menorData, maiorData, delta = tratadatas(inicioDesenv)
                        projetos.update({'Real Início do Desenv.': maiorData})
                        projetos.update({'Delta Início do Desenv.': delta})
                        
                        menorData, maiorData, delta = tratadatas(terminoDesenv)
                        projetos.update({'Real Término do Desenv.': maiorData})
                        projetos.update({'Delta Término do Desenv.': delta})
                        projetos.update({'Complitude Desenv': str(ws_Projetos.cell(row=linha, column=cfg.excelColumn['Complitude Desenv']).value)})
                        
                        inicioSmokeTest = convertDataParse(str(ws_Projetos.cell(row=linha, column=cfg.excelColumn['Início Smoke Test']).value))
                        projetos.update({'Início Smoke Test': inicioSmokeTest})
                        terminoSmokeTest = convertDataParse(str(ws_Projetos.cell(row=linha, column=cfg.excelColumn['Término Smoke Test']).value))
                        projetos.update({'Término Smoke Test': terminoSmokeTest})
                        projetos.update({'Complitude Test': str(ws_Projetos.cell(row=linha, column=cfg.excelColumn['Complitude Test']).value)})
                        
                        menorData, maiorData, delta = tratadatas(inicioSmokeTest)
                        projetos.update({'Real Início Smoke Test':maiorData})
                        projetos.update({'Delta Início Smoke Test':delta})

                        menorData, maiorData, delta = tratadatas(terminoSmokeTest)
                        projetos.update({'Real Término Smoke Test': maiorData})
                        projetos.update({'Delta Término Smoke Test': delta})


                        inicioTestesIntegrados = convertDataParse(str(ws_Projetos.cell(row=linha, column=cfg.excelColumn['Início Testes Integrados']).value))
                        projetos.update({'Início Testes Integrados': inicioTestesIntegrados})
                        terminoTestesIntegrados = convertDataParse(str(ws_Projetos.cell(row=linha, column=cfg.excelColumn['Término Testes Integrados']).value))
                        projetos.update({'Término Testes Integrados': terminoTestesIntegrados})
                        projetos.update({'Complitude Integrados': str(ws_Projetos.cell(row=linha, column=cfg.excelColumn['Complitude Integrados']).value)})
                        
                        menorData, maiorData, delta = tratadatas(inicioTestesIntegrados)
                        projetos.update({'Real Início Testes Integrados': maiorData})
                        projetos.update({'Delta Início Testes Integrados': delta})

                        menorData, maiorData, delta = tratadatas(terminoTestesIntegrados)
                        projetos.update({'Real Término Testes Integrados': maiorData})
                        projetos.update({'Delta Término Testes Integrados': delta})


                        inicioHomologacao = convertDataParse(str(ws_Projetos.cell(row=linha, column=cfg.excelColumn['Início da HML']).value))
                        projetos.update({'Início da HML': inicioHomologacao})
                        terminoHomologacao = convertDataParse(str(ws_Projetos.cell(row=linha, column=cfg.excelColumn['Término da HML']).value))
                        projetos.update({'Término da HML': terminoHomologacao})
                        projetos.update({'Complitude HML': str(ws_Projetos.cell(row=linha, column=cfg.excelColumn['Complitude HML']).value)})
                        
                        menorData, maiorData, delta = tratadatas(inicioHomologacao)
                        projetos.update({'Real Início da HML': maiorData})
                        projetos.update({'Delta Início da HML': delta})

                        menorData, maiorData, delta = tratadatas(terminoHomologacao)
                        projetos.update({'Real Término da HML': maiorData})
                        projetos.update({'Delta Término da HML': delta})

                        dataImplantacaoInicial = convertDataParse(str(ws_Projetos.cell(row=linha, column=cfg.excelColumn['Implantação (Go Live) (Início)']).value))
                        projetos.update({'Implantação (Go Live) (Início)': dataImplantacaoInicial})
                        menorData, maiorData, delta = tratadatas(dataImplantacaoInicial)
                        projetos.update({'Real Implantação (Go Live) (Início)': maiorData})
                        projetos.update({'Delta Implantação (Go Live) (Início)': delta})

                        projetos.update({'Implantação (Go Live) (Término)': convertDataParse(str(ws_Projetos.cell(row=linha, column=cfg.excelColumn['Implantação (Go Live) (Término)']).value))})
                        projetos.update({'Complitude Implantação': str(ws_Projetos.cell(row=linha, column=cfg.excelColumn['Complitude Implantação']).value)})
                        
                        projetos.update({'Descrição': ws_Projetos.cell(row=linha, column=cfg.excelColumn['Descrição']).value})
                        projetos.update({'A.N.': ws_Projetos.cell(row=linha, column=cfg.excelColumn['A.N.']).value})
                        projetos.update({'GP': ws_Projetos.cell(row=linha, column=cfg.excelColumn['GP']).value})
                        projetos.update({'LT': ws_Projetos.cell(row=linha, column=cfg.excelColumn['LT']).value})
                        projetos.update({'Líder de Testes': ws_Projetos.cell(row=linha, column=cfg.excelColumn['Líder de Testes']).value})
                        projetos.update({'Gerente Responsável': ws_Projetos.cell(row=linha, column=cfg.excelColumn['Gerente Responsável']).value})
                        projetos.update({'Patrocinador do Projeto (GD em INFRA)': ws_Projetos.cell(row=linha, column=cfg.excelColumn['Patrocinador do Projeto (GD em INFRA)']).value})
                        projetos.update({'Data de Término Planejada': convertDataParse(str(ws_Projetos.cell(row=linha, column=cfg.excelColumn['Data de Término Planejada']).value))})
                        projetos.update({'Gestão de Demandas': ws_Projetos.cell(row=linha, column=cfg.excelColumn['Gestão de Demandas']).value})
                        projetos.update({'Prioridade': ws_Projetos.cell(row=linha, column=cfg.excelColumn['Prioridade']).value})
                        projetos.update({'Objetivo - Macro': ws_Projetos.cell(row=linha, column=cfg.excelColumn['Objetivo - Macro']).value})
                        projetos.update({'Funding': ws_Projetos.cell(row=linha, column=cfg.excelColumn['Funding']).value})
                        projetos.update({'Solicitante': ws_Projetos.cell(row=linha, column=cfg.excelColumn['Solicitante']).value})
                        projetos.update({'Project Name': ws_Projetos.cell(row=linha, column=cfg.excelColumn['Project Name']).value})
                        projetos.update({'Fórmula Status Fase': ws_Projetos.cell(row=linha, column=cfg.excelColumn['Fórmula Status Fase']).value})
                        projetos.update({'Fórmula Status Projeto': ws_Projetos.cell(row=linha, column=cfg.excelColumn['Fórmula Status Projeto']).value})
                        projetos.update({'Forecast Status da Entrega': ws_Projetos.cell(row=linha, column=cfg.excelColumn['Forecast Status da Entrega']).value})
                        listaDeProjetos.append(projetos)
    salvaProjetos(listaDeProjetos)

def creation_date(path_to_file):
    """
    Try to get the date that a file was created, falling back to when it was
    last modified if that isn't possible.
    See http://stackoverflow.com/a/39501288/1709587 for explanation.
    """
    if platform.system() == 'Windows':
        return os.path.getctime(path_to_file)
    else:
        stat = os.stat(path_to_file)
        try:
            return stat.st_birthtime
        except AttributeError:
            # We're probably on Linux. No easy way to get creation dates here,
            # so we'll settle for when its content was last modified.
            return stat.st_mtime


def main():
    readFilesFromSource()
    # print(cfg.excelColumn['Implantação (Go Live) (Início)'])


if __name__ == "__main__":
	main()